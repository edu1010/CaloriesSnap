#!/usr/bin/env python3
"""Import AESAN product nutrition data from XLSX into app JSON format."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ROW = 2
DATA_START_ROW = 3
MIN_BARCODE_LEN = 8
MAX_BARCODE_LEN = 14


def _normalize_barcode(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None

    if raw.endswith(".0"):
        raw = raw[:-2]

    digits = re.sub(r"\D", "", raw)
    if not (MIN_BARCODE_LEN <= len(digits) <= MAX_BARCODE_LEN):
        return None
    return digits


def _as_clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        return float(match.group(0))
    except ValueError:
        return None


def _build_product_name(name: str | None, brand: str | None) -> str | None:
    if name is None:
        return None
    if brand is None:
        return name
    if name.upper().startswith(brand.upper()):
        return name
    return f"{name} ({brand})"


def parse_aesan_xlsx(input_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if "DATOS" not in workbook.sheetnames:
        raise ValueError("Sheet 'DATOS' not found in workbook.")

    ws = workbook["DATOS"]
    headers = [c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
    index = {name: i for i, name in enumerate(headers) if name}

    required = [
        "EAN",
        "Nombre comercial",
        "Marca",
        "Categoría",
        "Subcategoría",
        "Fuente",
        "Energía \n(kCal/ 100g ó 100 ml)",
    ]
    missing = [key for key in required if key not in index]
    if missing:
        raise ValueError(f"Missing required columns in DATOS sheet: {missing}")

    by_barcode: dict[str, dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        barcode = _normalize_barcode(row[index["EAN"]])
        kcal = _parse_float(row[index["Energía \n(kCal/ 100g ó 100 ml)"]])
        if barcode is None or kcal is None or kcal <= 0:
            continue

        brand = _as_clean_text(row[index["Marca"]])
        commercial_name = _as_clean_text(row[index["Nombre comercial"]])
        name = _build_product_name(commercial_name, brand)
        if name is None:
            continue

        category = _as_clean_text(row[index["Categoría"]])
        subcategory = _as_clean_text(row[index["Subcategoría"]])
        source = _as_clean_text(row[index["Fuente"]])

        current = by_barcode.get(barcode)
        candidate = {
            "name": name,
            "kcalPer100g": round(kcal, 2),
            "defaultGramsSmall": 50,
            "defaultGramsMedium": 100,
            "defaultGramsLarge": 150,
            "barcode": barcode,
            "brand": brand,
            "category": " > ".join(
                [part for part in (category, subcategory) if part],
            )
            or None,
            "source": f"AESAN 2022 ({source})" if source else "AESAN 2022",
        }

        if current is None:
            by_barcode[barcode] = candidate
            continue

        # Prefer richer names when duplicated barcodes are present.
        if len(candidate["name"]) > len(current["name"]):
            by_barcode[barcode] = candidate

    products = sorted(
        by_barcode.values(),
        key=lambda item: (item.get("name") or "", item["barcode"]),
    )
    return products


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert AESAN XLSX into CalorieSnap nutrition product JSON.",
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to AESAN XLSX file (e.g. BasedatosWeb.xlsx).",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Destination JSON path.",
    )
    args = parser.parse_args()

    products = parse_aesan_xlsx(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(products, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Written {len(products)} products to {args.output}")


if __name__ == "__main__":
    main()
